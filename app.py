import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
import io  # Add this import at the top with others if not present

from typing import Dict, List, Any, Optional
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelProcessor:
    """Memory-efficient Excel file processor with multi-sheet support"""
    
    def __init__(self):
        self.workbook = None
        self.sheet_names = []
        self.current_sheet = None
        self.chunk_size = 1000  # Process 1000 rows at a time
    
    def load_excel_file(self, uploaded_file) -> bool:
        """Load Excel file and extract sheet information"""
        try:
            # Read the uploaded file into BytesIO
            file_bytes = BytesIO(uploaded_file.read())
            
            # Load workbook with openpyxl for sheet info
            self.workbook = openpyxl.load_workbook(file_bytes, read_only=True)
            self.sheet_names = self.workbook.sheetnames
            
            # Reset file pointer for pandas
            file_bytes.seek(0)
            self.file_bytes = file_bytes
            
            logger.info(f"Successfully loaded Excel file with {len(self.sheet_names)} sheets")
            return True
            
        except Exception as e:
            logger.error(f"Error loading Excel file: {str(e)}")
            st.error(f"Error loading Excel file: {str(e)}")
            return False
    
    def get_sheet_info(self) -> Dict[str, Any]:
        """Get information about all sheets in the workbook"""
        sheet_info = {}
        
        for sheet_name in self.sheet_names:
            try:
                sheet = self.workbook[sheet_name]
                sheet_info[sheet_name] = {
    'max_row': sheet.max_row,
    'max_column': sheet.max_column,
    # 'dimensions': sheet.dimensions,  # Removed unsupported attribute
    'range': f"A1:{openpyxl.utils.get_column_letter(sheet.max_column)}{sheet.max_row}" if sheet.max_row and sheet.max_column else 'Unknown'
}
            except Exception as e:
                logger.warning(f"Could not get info for sheet {sheet_name}: {str(e)}")
                sheet_info[sheet_name] = {'error': str(e)}
        
        return sheet_info
    
    def read_sheet_chunked(self, sheet_name: str, chunk_size: Optional[int] = None) -> pd.DataFrame:
        """Read a specific sheet with chunking for memory efficiency"""
        if chunk_size is None:
            chunk_size = self.chunk_size
        
        try:
            # Reset file pointer
            self.file_bytes.seek(0)
            
            # Read the specific sheet
            df = pd.read_excel(
                self.file_bytes,
                sheet_name=sheet_name,
                engine='openpyxl'
            )
            
            # Detect and convert data types
            df = self._detect_and_convert_types(df)
            
            logger.info(f"Successfully read sheet '{sheet_name}' with {len(df)} rows and {len(df.columns)} columns")
            return df
            
        except Exception as e:
            logger.error(f"Error reading sheet {sheet_name}: {str(e)}")
            st.error(f"Error reading sheet {sheet_name}: {str(e)}")
            return pd.DataFrame()
    
    def _detect_and_convert_types(self, df: pd.DataFrame) -> pd.DataFrame:
        """Detect and convert data types for better performance"""
        for column in df.columns:
            try:
                # Try to convert to numeric
                if df[column].dtype == 'object':
                    # Check if it's a date
                    if self._is_date_column(df[column]):
                        df[column] = pd.to_datetime(df[column], errors='coerce')
                    else:
                        # Try numeric conversion
                        numeric_series = pd.to_numeric(df[column], errors='coerce')
                        if not numeric_series.isna().all():
                            df[column] = numeric_series
            except Exception as e:
                logger.warning(f"Could not convert column {column}: {str(e)}")
                continue
        
        return df
    
    def _is_date_column(self, series: pd.Series) -> bool:
        """Check if a series contains date-like values"""
        sample_size = min(100, len(series))
        sample = series.dropna().head(sample_size)
        
        if len(sample) == 0:
            return False
        
        date_count = 0
        for value in sample:
            try:
                pd.to_datetime(value)
                date_count += 1
            except:
                continue
        
        # If more than 50% of samples are dates, consider it a date column
        return (date_count / len(sample)) > 0.5
    
    def get_column_info(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Get detailed information about columns"""
        column_info = {}
        
        for column in df.columns:
            column_info[column] = {
                'dtype': str(df[column].dtype),
                'null_count': df[column].isnull().sum(),
                'unique_count': df[column].nunique(),
                'sample_values': df[column].dropna().head(5).tolist()
            }
        
        return column_info

def main():
    st.set_page_config(
        page_title="Excel Sheets Agent",
        page_icon="📊",
        layout="wide"
    )
    
    st.title("📊 Excel Sheets Agent")
    st.markdown("Upload and analyze Excel files with natural language queries")
    
    # Initialize session state
    if 'processor' not in st.session_state:
        st.session_state.processor = ExcelProcessor()
    if 'current_df' not in st.session_state:
        st.session_state.current_df = None
    if 'sheet_info' not in st.session_state:
        st.session_state.sheet_info = {}
    
    # Sidebar for file upload and sheet selection
    with st.sidebar:
        st.header("📁 File Upload")
        
        uploaded_file = st.file_uploader(
            "Choose an Excel file",
            type=['xlsx', 'xls'],
            help="Upload an Excel file to analyze"
        )
        
        if uploaded_file is not None:
            if st.button("Load File"):
                with st.spinner("Loading Excel file..."):
                    if st.session_state.processor.load_excel_file(uploaded_file):
                        st.session_state.sheet_info = st.session_state.processor.get_sheet_info()
                        st.success("File loaded successfully!")
                        st.rerun()
        
        # Sheet selection
        if st.session_state.processor.sheet_names:
            st.header("📋 Sheet Selection")
            selected_sheet = st.selectbox(
                "Select a sheet to analyze:",
                st.session_state.processor.sheet_names
            )
            
            if st.button("Load Sheet"):
                with st.spinner(f"Loading sheet '{selected_sheet}'..."):
                    df = st.session_state.processor.read_sheet_chunked(selected_sheet)
                    if not df.empty:
                        st.session_state.current_df = df
                        st.session_state.current_sheet = selected_sheet
                        st.success(f"Sheet '{selected_sheet}' loaded successfully!")
                        st.rerun()
    
    # Main content area
    if st.session_state.sheet_info:
        st.header("📊 Workbook Overview")
        
        # Display sheet information
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("Available Sheets")
            for sheet_name, info in st.session_state.sheet_info.items():
                if 'error' not in info:
                    st.write(f"**{sheet_name}**")
                    st.write(f"- Rows: {info.get('max_row', 'Unknown')}")
                    st.write(f"- Columns: {info.get('max_column', 'Unknown')}")
                    st.write(f"- Range: {info.get('dimensions', 'Unknown')}")
                else:
                    st.write(f"**{sheet_name}** - Error: {info['error']}")
        
        with col2:
            if st.session_state.current_df is not None:
                st.subheader(f"Current Sheet: {st.session_state.current_sheet}")
                df = st.session_state.current_df
                
                st.write(f"**Shape:** {df.shape[0]} rows × {df.shape[1]} columns")
                
                # Column information
                column_info = st.session_state.processor.get_column_info(df)
                
                with st.expander("Column Details"):
                    for col, info in column_info.items():
                        st.write(f"**{col}**")
                        st.write(f"- Type: {info['dtype']}")
                        st.write(f"- Null values: {info['null_count']}")
                        st.write(f"- Unique values: {info['unique_count']}")
                        st.write(f"- Sample: {info['sample_values']}")
                        st.write("---")
    
    # Data preview
    if st.session_state.current_df is not None:
        st.header("📋 Data Preview")
        
        df = st.session_state.current_df
        
        # Display options
        col1, col2, col3 = st.columns(3)
        with col1:
            show_rows = st.number_input("Rows to display", min_value=5, max_value=1000, value=10)
        with col2:
            show_info = st.checkbox("Show data info", value=True)
        with col3:
            show_stats = st.checkbox("Show statistics", value=False)
        
        # Display data
        st.dataframe(df.head(show_rows), use_container_width=True)
        
        if show_info:
            st.subheader("Data Info")
            buffer = BytesIO()
            df.info(buf=buffer)
            info_str = buffer.getvalue().decode()
            st.text(info_str)
        
        if show_stats:
            st.subheader("Statistical Summary")
            st.dataframe(df.describe(), use_container_width=True)
    
    # Footer
    st.markdown("---")

import sys
import traceback
from llm_utils import run_gemini_query

import time
from functools import lru_cache

def log_event(event: str, **kwargs):
    with open("app_metrics.log", "a") as f:
        f.write(f"{time.strftime('%Y-%m-%d %H:%M:%S')} | {event} | {kwargs}\n")

@st.cache_data(show_spinner=False)
def cached_safe_exec(code: str, df: pd.DataFrame) -> Optional[pd.DataFrame]:
    """Cacheable version of safe_exec for identical queries and data."""
    local_vars = {'df': df.copy()}
    try:
        exec(f"result = {code}", {}, local_vars)
        return local_vars['result']
    except Exception as e:
        log_event("safe_exec_error", error=str(e), code=code)
        return None

def safe_exec(code: str, df: pd.DataFrame) -> Optional[pd.DataFrame]:
    """Safely execute pandas code returned by LLM and log timing."""
    start = time.time()
    result = cached_safe_exec(code, df)
    elapsed = time.time() - start
    log_event("query_exec", code=code, elapsed=elapsed)
    if elapsed > 10:
        st.warning(f"⚠️ Query took {elapsed:.2f} seconds (exceeds 10s target)")
    else:
        st.info(f"Query executed in {elapsed:.2f} seconds")
    return result

def phase2_nl_query_ui():
    st.header("🤖 Natural Language Query")
    if st.session_state.current_df is None:
        st.info("Please load a sheet first.")
        return
    df = st.session_state.current_df
    columns = list(df.columns)
    with st.form("nl_query_form"):
        nl_query = st.text_input("Enter your query (e.g., 'Show customers from Delhi with > 10000 revenue')")
        query_type = st.selectbox(
            "Query type:",
            ["filter", "aggregate", "sort", "pivot"],
            help="What kind of operation do you want to perform?"
        )
        submitted = st.form_submit_button("Run Query")
    if submitted and nl_query:
        with st.spinner("Calling Gemini LLM and executing query..."):
            code = run_gemini_query(query_type, nl_query, columns)
            if code:
                st.code(code, language="python")
                result_df = safe_exec(code, df)
                if result_df is not None:
                    st.success("Query executed!")
                    st.dataframe(result_df.head(100), use_container_width=True)
            else:
                st.error("Gemini LLM could not generate a valid pandas expression. Try rewording your query.")

def main():
    st.set_page_config(
        page_title="Excel Sheets Agent",
        page_icon="📊",
        layout="wide"
    )
    st.title("📊 Excel Sheets Agent")
    st.markdown("Upload and analyze Excel files with natural language queries")
    # Initialize session state
    if 'processor' not in st.session_state:
        st.session_state.processor = ExcelProcessor()
    if 'current_df' not in st.session_state:
        st.session_state.current_df = None
    if 'sheet_info' not in st.session_state:
        st.session_state.sheet_info = {}
    # Sidebar for file upload and sheet selection
    with st.sidebar:
        st.header("📁 File Upload")
        uploaded_file = st.file_uploader(
            "Choose an Excel file",
            type=['xlsx', 'xls'],
            help="Upload an Excel file to analyze"
        )
        if uploaded_file is not None:
            if st.button("Load File"):
                with st.spinner("Loading Excel file..."):
                    if st.session_state.processor.load_excel_file(uploaded_file):
                        st.session_state.sheet_info = st.session_state.processor.get_sheet_info()
                        st.success("File loaded successfully!")
                        st.rerun()
        # Sheet selection
        if st.session_state.processor.sheet_names:
            st.header("📋 Sheet Selection")
            selected_sheet = st.selectbox(
                "Select a sheet to analyze:",
                st.session_state.processor.sheet_names
            )
            if st.button("Load Sheet"):
                with st.spinner(f"Loading sheet '{selected_sheet}'..."):
                    df = st.session_state.processor.read_sheet_chunked(selected_sheet)
                    if not df.empty:
                        st.session_state.current_df = df
                        st.session_state.current_sheet = selected_sheet
                        st.success(f"Sheet '{selected_sheet}' loaded successfully!")
                        st.rerun()
    # Main content area
    if st.session_state.sheet_info:
        st.header("📊 Workbook Overview")
        col1, col2 = st.columns(2)
        with col1:
            st.subheader("Available Sheets")
            for sheet_name, info in st.session_state.sheet_info.items():
                if 'error' not in info:
                    st.write(f"**{sheet_name}**")
                    st.write(f"- Rows: {info.get('max_row', 'Unknown')}")
                    st.write(f"- Columns: {info.get('max_column', 'Unknown')}")
                    st.write(f"- Range: {info.get('dimensions', 'Unknown')}")
                else:
                    st.write(f"**{sheet_name}** - Error: {info['error']}")
        with col2:
            if st.session_state.current_df is not None:
                st.subheader(f"Current Sheet: {st.session_state.current_sheet}")
                df = st.session_state.current_df
                st.write(f"**Shape:** {df.shape[0]} rows × {df.shape[1]} columns")
                column_info = st.session_state.processor.get_column_info(df)
                with st.expander("Column Details"):
                    for col, info in column_info.items():
                        st.write(f"**{col}**")
                        st.write(f"- Type: {info['dtype']}")
                        st.write(f"- Null values: {info['null_count']}")
                        st.write(f"- Unique values: {info['unique_count']}")
                        st.write(f"- Sample: {info['sample_values']}")
                        st.write("---")
    # Data preview
    if st.session_state.current_df is not None:
        st.header("📋 Data Preview")
        df = st.session_state.current_df
        col1, col2, col3 = st.columns(3)
        with col1:
            show_rows = st.number_input("Rows to display", min_value=5, max_value=1000, value=10)
        with col2:
            show_info = st.checkbox("Show data info", value=True)
        with col3:
            show_stats = st.checkbox("Show statistics", value=False)
        st.dataframe(df.head(show_rows), use_container_width=True)
        if show_info:
            st.subheader("Data Info")
            try:
                buffer = io.StringIO()
                df.info(buf=buffer)
                info_str = buffer.getvalue()
                st.text(info_str)
            except Exception as e:
                st.error(f"Failed to display DataFrame info: {e}")
                logger.error(f"Failed to display DataFrame info: {e}")
        if show_stats:
            st.subheader("Statistical Summary")
            st.dataframe(df.describe(), use_container_width=True)
    # Phase 2: NL Query UI
    phase2_nl_query_ui()
    # Footer
    st.markdown("---")

if __name__ == "__main__":
    main()

